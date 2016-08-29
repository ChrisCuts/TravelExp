'''
Created on 26.07.2016

@author: DoppC
'''

import PyPDF2
from openpyxl import load_workbook
import configparser

from datetime import datetime
import os
from shutil import copyfile
import re

import sys

#from openpyxl import load_workbook

class DBTrip():
    
    class UmlautDecode():
        TABLE = [(str(b'\xe2\x80\xa6', encoding='utf-8'), u'ö'),
                 (str(b'\xe2\x80\x94', encoding='utf-8'), u'ö'),
                 (str(b'\xe2\x80\x93', encoding='utf-8'), u'ö'),
                 (str(b'\xe2\x80\xa0', encoding='utf-8'), u'ü'),
                 (str(b'\xe2\x80\xa2', encoding='utf-8'), u'ä')]
    
        @classmethod
        def decode(cls, lines):
            for code, umlaut in cls.TABLE:
                lines = lines.replace(code, umlaut)
            return lines
        
    class Spacetime():
        
        def __init__(self, space, date, HHMM):
            
            self.time = datetime.strptime(date + HHMM, '%d.%m.%Y%H:%M')
            self.space = space
            self.date = date
            
        def __str__(self):
            
            return '{0:%H}:{0:%M}\t{1}'.format(self.time, self.space)
        
        def week(self):
            
            return self.time.isocalendar()[1]
            
    def __init__(self, path, filename):
        
        
        file = open(path + '/' + filename, 'rb')
    
        pdf = PyPDF2.PdfFileReader(file)
        page = pdf.getPage(0)
        content = PyPDF2.pdf.ContentStream(page["/Contents"].getObject(), page.pdf)
        
        lines = ''
        for operands, operator in content.operations:
            if operator == PyPDF2.pdf.utils.b_("Tj"):
                lines += operands[0] + '\n'
            elif operator == PyPDF2.pdf.utils.b_("TJ"):
                for i in operands[0]:
                    if isinstance(i, str):
                        lines += i
                lines += '\n'
        
        lines = self.UmlautDecode.decode(lines)
                
        year = re.search('(?<=G.ltigkeit: \n\d\d.\d\d.)\d{4}', lines).group()
        
        self.costs = float(re.search('(?<=Summe\n)\d\d(?=,\d\d)', lines).group()
                           + '.'
                           + re.search('(?<=Summe\n\d\d,)\d\d', lines).group())
        
        date = re.search('(?<=\n)\d\d.\d\d.(?=\nab)', lines).group()
        time = re.search('(?<=\nab )\d\d:\d\d(?=\n)', lines).group()
        space = re.search('(?<=\n)[A-ZÄÖÜa-zäöü0-9, ]*(?=\n\d\d.\d\d.\nab)', lines).group()
        self.departure = self.Spacetime(space, date + year, time)        
        
        time = re.findall('(?<=\nan )\d\d:\d\d(?=\n)', lines)[-1]
        space = re.findall('(?<=\n)[A-ZÄÖÜa-zäöü0-9, ]*(?=\n\d\d.\d\d.\nan)', lines)[-1]
        self.arrival = self.Spacetime(space, date + year, time)
        
        file.close()
        
        #rename file
        try:
            os.rename(path + '/' + filename, path + '/' + str(self.departure.time.day) + os.path.splitext(filename)[1])
        except PermissionError:
            pass
        
    def __str__(self):
        
        res  = 'Date:\t\t' + self.arrival.date
        res += '\n\nDeparture:\t' + self.departure.__str__()
        res += '\nArrival:\t' + self.arrival.__str__()
        res += '\n\nCosts:\t\t' + self.costs
        return res

    def __lt__(self, other):
        
        return self.departure.time < other.arrival.time
        
        
class TravelExpense():
    
    class Person:
        def __init__(self, config):
            self.name   = config.get('PERSON', 'name')
            self.id     = config.get('PERSON', 'id')
        
    class Project:
        def __init__(self, config):
            self.name   = config.get('PROJECT', 'name')
            self.id     = config.get('PROJECT', 'id')
    
    class Hotel:
        def __init__(self, config):
            self.name   = config.get('HOTEL', 'name')
            self.breakfast = config.get('HOTEL', 'breakfast')
            self.costs  = config.getfloat('HOTEL', 'costs')
    
    class Expense:
        def __init__(self, trips, week):
            self.date_of_issue = datetime.today()
            self.number = '{:02n}{:02n}'.format(trips[0].departure.time.month, week)
        
    class Trips:
        def __init__(self, trips):
            # TRIPS
            if len(trips) > 2:
                raise(ValueError('Only two trip allowed in a week'))
            elif len(trips) == 2:
                
                
                self.dates = [trips[0].departure, trips[1].arrival]
                self.costs = [trips[0].costs, trips[1].costs]
                self.nights = (trips[1].arrival.time.date() - trips[0].departure.time.date()).days
                
                self.route = [trips[0].departure.space + ' - ' + trips[0].arrival.space,
                              trips[1].departure.space + ' - ' + trips[1].arrival.space]
                
            elif len(trips) == 1:
                
                
                self.nights = None
                
                if trips[0].departure.time.weekday() + trips[0].departure.time.hour / 24 <= 2.5:
                    self.dates = [trips[0].departure, None]
                    self.costs = [trips[0].costs, None]
                    self.route = [trips[0].departure.space + ' - ' + trips[0].arrival.space, None]
                
                elif trips[0].departure.time.weekday() + trips[0].departure.time.hour / 24 > 2.5:
                    self.dates = [None, trips[0].arrival]
                    self.costs = [None, trips[0].costs]
                    self.route = [None, trips[0].departure.space + ' - ' + trips[0].arrival.space]
                
            else:
                raise(ValueError('Invalid argument'))
            
    # PERSON
    NAME        = 'B4'
    ID          = 'I4'
    
    # PROJECT
    PROJ_NAME   = 'B5'
    PROJ_ID     = 'E6'
    
    # HOTEL
    HOTEL_NAME  = 'B12'
    HOTEL_BREAKFAST = 'B11'
    HOTEL_COSTS = 'I12'
    
    # EXPENSE
    DATE_OF_ISSUE = 'B3'
    EXP_NUMBER      = 'F3'
    
    # TRIP
    START_DATE  = 'C8'
    START_TIME  = 'G8'
    END_DATE    = 'C9'
    END_TIME    = 'G9'
    
    START_ROUTE = 'B23'
    START_COSTS = 'H23'
    END_ROUTE   = 'B24'
    END_COSTS   = 'H24'
    
    
    def __init__(self, week, trips, path, config):
        
        self.path = path
        week = week
        
        self.person     = self.Person(config)
        self.project    = self.Project(config)
        self.hotel      = self.Hotel(config)
        
        self.expense    = self.Expense(trips, week)
        self.trips      = self.Trips(trips)
        

        self.edit_xls()
        
    def edit_xls(self):
        
        self.file = '{}/Reisekosten_{}.xlsx'.format(self.path, self.expense.number)
        copyfile('Reisekostenabrechnung.xlsx', self.file)
        
        self.fill_cells()    
        
    def fill_cells(self):
        
        wb = load_workbook(self.file)
        ws = wb.active
        
        # PERSON
        ws[self.NAME]   = self.person.name
        ws[self.ID]     = self.person.id
        
        # PROJECT
        ws[self.PROJ_NAME]  = self.project.name
        ws[self.PROJ_ID]    = self.project.id
        
        # HOTEL
        ws[self.HOTEL_NAME]         = self.hotel.name
        ws[self.HOTEL_BREAKFAST]    = self.hotel.breakfast
        if self.trips.nights != None:
            ws[self.HOTEL_COSTS]    = self.hotel.costs * self.trips.nights
        
        # EXPENSE
        ws[self.DATE_OF_ISSUE]  = self.expense.date_of_issue
        ws[self.DATE_OF_ISSUE].number_format = 'DD.MM.YYYY'
            
        ws[self.EXP_NUMBER]         = self.expense.number
        
        # TRIP
        if self.trips.dates[0] == None:
            ws[self.START_DATE] = None
            ws[self.START_DATE].number_format = 'DD.MM.YYYY'
            ws[self.START_TIME] = None
            ws[self.START_TIME].number_format = 'HH:MM'
        
            ws[self.START_ROUTE] = None
            ws[self.START_COSTS]= None
        else:
            ws[self.START_DATE] = self.trips.dates[0].time.date()
            ws[self.START_DATE].number_format = 'DD.MM.YYYY'
            ws[self.START_TIME] = self.trips.dates[0].time.time()
            ws[self.START_TIME].number_format = 'HH:MM'
        
            ws[self.START_ROUTE] = self.trips.route[0]
            ws[self.START_COSTS]= self.trips.costs[0]
        
        if self.trips.dates[1] == None:
            ws[self.END_DATE]   = None
            ws[self.END_DATE].number_format = 'DD.MM.YYYY'
            ws[self.END_TIME]   = None
            ws[self.END_TIME].number_format = 'HH:MM'
        
            ws[self.END_ROUTE]  = None
            ws[self.END_COSTS]  = None
        else:
            ws[self.END_DATE]   = self.trips.dates[1].time.date()
            ws[self.END_DATE].number_format = 'DD.MM.YYYY'
            ws[self.END_TIME]   = self.trips.dates[1].time.time()
            ws[self.END_TIME].number_format = 'HH:MM'
        
            ws[self.END_ROUTE]  = self.trips.route[1]
            ws[self.END_COSTS]  = self.trips.costs[1]
            
            
        wb.save(self.file)
        
        
class TicketFolder():
    
    def __init__(self, path):
    
        # scan config
        config = configparser.RawConfigParser()
        config.read('expense.cfg')
    
        # scan folder    
        ticketfiles = os.listdir(path)        
        
        tickets = []
        for ticketfile in ticketfiles:
            if ticketfile.endswith('.pdf'):
                try:
                    tickets.append(DBTrip(path, ticketfile))
                except:
                    print(ticketfile + '\t failed')
#                     print(sys.exc_info()[0])
#                     print(sys.exc_info()[1])
                    raise
        tickets.sort()
        
        
        # create expenses
        self.expenses = []
        weekly = []
        for ticket in tickets:
            
            if weekly == [] or ticket.departure.week() == weekly[0].departure.week():
                weekly.append(ticket)
                
            else:
                
                self.expenses.append(TravelExpense(len(self.expenses)+1, weekly, path, config))
                weekly.clear()
                weekly.append(ticket)
            
        self.expenses.append(TravelExpense(len(self.expenses)+1, weekly, path, config))
        
        


if __name__ == '__main__':
    
    if len(sys.argv) != 2:
        print('Creates a travel expense by analysing DB Ticket documents.')
        print('\nUsage: travelexp.exe <path to documents>')
    else:
        
        folder = TicketFolder(sys.argv[1])
    
#    pdf = PyPDF2.PdfFileReader('August/15.pdf')
#    page = pdf.getPage(0)
#    content = page["/Contents"].getObject()
#    content = PyPDF2.pdf.ContentStream(content, page.pdf)
#    
#     trip = DBTrip('August', '11.pdf')
#     print(trip)
#    wb = load_workbook('Mappe1.xlsx')
#    
#    ws = wb.active
#    ws['A3'] = 42
#    wb.save("Mappe1.xlsx")